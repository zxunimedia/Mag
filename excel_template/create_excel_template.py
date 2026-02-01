#!/usr/bin/env python3
"""
文化部原村管考系統 - Excel 匯入架構範本
包含三個分頁：計畫資料、月報表、經費支出
具備下拉選單連動與 VLOOKUP 公式
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule

# 主題色彩 - 使用文化部風格
THEME = {
    'primary': '2D3E50',      # 深藍灰
    'accent': 'E67E22',       # 橘色強調
    'light': 'ECF0F1',        # 淺灰背景
    'success': '27AE60',      # 綠色
    'warning': 'F39C12',      # 警告黃
    'danger': 'E74C3C',       # 紅色
    'text': '2C3E50',         # 文字色
    'border': 'BDC3C7',       # 邊框色
}

# 字型設定
TITLE_FONT = Font(name='Microsoft JhengHei', size=16, bold=True, color=THEME['primary'])
HEADER_FONT = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Microsoft JhengHei', size=10, color=THEME['text'])
LINK_FONT = Font(name='Microsoft JhengHei', size=10, color=THEME['accent'], underline='single')

# 填充樣式
HEADER_FILL = PatternFill(start_color=THEME['primary'], end_color=THEME['primary'], fill_type='solid')
ALT_FILL = PatternFill(start_color=THEME['light'], end_color=THEME['light'], fill_type='solid')

# 邊框樣式
THIN_BORDER = Border(
    left=Side(style='thin', color=THEME['border']),
    right=Side(style='thin', color=THEME['border']),
    top=Side(style='thin', color=THEME['border']),
    bottom=Side(style='thin', color=THEME['border'])
)

# 對齊樣式
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


def create_workbook():
    """建立工作簿"""
    wb = Workbook()
    
    # 建立三個分頁
    ws1 = wb.active
    ws1.title = "計畫資料"
    ws2 = wb.create_sheet("月報表")
    ws3 = wb.create_sheet("經費支出")
    
    # 設定各分頁
    setup_project_data_sheet(ws1)
    setup_monthly_report_sheet(ws2)
    setup_expenditure_sheet(ws3)
    
    return wb


def setup_project_data_sheet(ws):
    """分頁一：計畫資料（主表）"""
    ws.sheet_view.showGridLines = False
    
    # 標題
    ws.merge_cells('B2:G2')
    ws['B2'] = '計畫資料（主表）'
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 35
    
    # 說明文字
    ws.merge_cells('B3:G3')
    ws['B3'] = '此表定義計畫的關鍵項目與預算科目，供月報表與經費支出連動使用'
    ws['B3'].font = Font(name='Microsoft JhengHei', size=9, color='7F8C8D')
    
    # 表頭
    headers = ['關鍵項目 ID', '工作項目名稱', '預算科目', '核定金額', '預定目標值', '預計完成日']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 30
    
    # 範例資料
    sample_data = [
        ['KR-001', '辦理計畫說明會', '01-人事費', 50000, 1, '2026/03/31'],
        ['KR-002', '耆老訪談紀錄', '02-業務費', 80000, 10, '2026/06/30'],
        ['KR-003', '文化體驗活動', '02-業務費', 150000, 3, '2026/09/30'],
        ['KR-004', '成果發表會', '02-業務費', 100000, 1, '2026/12/15'],
        ['KR-005', '行政雜支', '03-雜支', 20000, 0, '2026/12/31'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=6):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN if col_idx in [2, 4, 5, 6, 7] else LEFT_ALIGN
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
            # 金額格式
            if col_idx == 5:
                cell.number_format = '#,##0'
                cell.alignment = RIGHT_ALIGN
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    
    # 定義名稱範圍（供下拉選單使用）
    # 關鍵項目 ID 範圍
    ws.parent.create_named_range('關鍵項目ID', ws, '$B$6:$B$20')
    # 預算科目範圍
    ws.parent.create_named_range('預算科目', ws, '$D$6:$D$20')


def setup_monthly_report_sheet(ws):
    """分頁二：月報表（填報）"""
    ws.sheet_view.showGridLines = False
    
    # 標題
    ws.merge_cells('B2:H2')
    ws['B2'] = '月報表（填報）'
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 35
    
    # 說明文字
    ws.merge_cells('B3:H3')
    ws['B3'] = '工作事項透過下拉選單連動「計畫資料」的關鍵項目 ID，選擇後自動帶出項目名稱與預定目標值'
    ws['B3'].font = Font(name='Microsoft JhengHei', size=9, color='7F8C8D')
    
    # 表頭
    headers = ['填報月份', '對應項目 ID', '工作項目名稱', '預定目標值', '本月達成數', '達成率(%)', '本月執行內容']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 30
    
    # 範例資料（含公式）
    sample_data = [
        ['2026年01月', 'KR-001', None, None, 1, None, '完成計畫說明會籌備工作'],
        ['2026年01月', 'KR-002', None, None, 3, None, '訪談 3 位部落耆老'],
        ['2026年02月', 'KR-002', None, None, 4, None, '訪談 4 位部落耆老'],
        ['2026年02月', 'KR-003', None, None, 0, None, '活動場地勘查'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=6):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx == 4:  # 工作項目名稱 - VLOOKUP
                cell.value = f'=IFERROR(VLOOKUP(C{row_idx},計畫資料!$B$6:$G$20,2,FALSE),"")'
            elif col_idx == 5:  # 預定目標值 - VLOOKUP
                cell.value = f'=IFERROR(VLOOKUP(C{row_idx},計畫資料!$B$6:$G$20,5,FALSE),"")'
            elif col_idx == 7:  # 達成率 - 計算公式
                cell.value = f'=IFERROR(F{row_idx}/E{row_idx},"")'
                cell.number_format = '0.0%'
            else:
                cell.value = value
            
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN if col_idx in [2, 3, 5, 6, 7] else LEFT_ALIGN
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
    
    # 設定下拉選單 - 對應項目 ID
    dv = DataValidation(
        type="list",
        formula1='=關鍵項目ID',
        allow_blank=True
    )
    dv.error = '請從下拉選單選擇有效的關鍵項目 ID'
    dv.errorTitle = '輸入錯誤'
    dv.prompt = '請選擇對應的關鍵項目 ID'
    dv.promptTitle = '選擇項目'
    ws.add_data_validation(dv)
    dv.add('C6:C100')
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35
    
    # 公式說明區
    ws.merge_cells('B12:H12')
    ws['B12'] = '【VLOOKUP 公式說明】'
    ws['B12'].font = Font(name='Microsoft JhengHei', size=11, bold=True, color=THEME['accent'])
    
    formulas_info = [
        ('工作項目名稱公式：', '=VLOOKUP(C6,計畫資料!$B$6:$G$20,2,FALSE)'),
        ('預定目標值公式：', '=VLOOKUP(C6,計畫資料!$B$6:$G$20,5,FALSE)'),
        ('達成率公式：', '=F6/E6 (本月達成數 ÷ 預定目標值)'),
    ]
    
    for idx, (label, formula) in enumerate(formulas_info, start=13):
        ws.cell(row=idx, column=2, value=label).font = Font(name='Microsoft JhengHei', size=9, bold=True)
        ws.cell(row=idx, column=3, value=formula).font = Font(name='Consolas', size=9, color='7F8C8D')


def setup_expenditure_sheet(ws):
    """分頁三：經費支出（核銷）"""
    ws.sheet_view.showGridLines = False
    
    # 標題
    ws.merge_cells('B2:I2')
    ws['B2'] = '經費支出（核銷）'
    ws['B2'].font = TITLE_FONT
    ws['B2'].alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 35
    
    # 說明文字
    ws.merge_cells('B3:I3')
    ws['B3'] = '支出項目透過下拉選單連動「計畫資料」的預算科目，選擇後自動帶出核定金額與剩餘餘額'
    ws['B3'].font = Font(name='Microsoft JhengHei', size=9, color='7F8C8D')
    
    # 表頭
    headers = ['日期', '對應科目', '科目名稱', '核定金額', '經費來源', '本月報支金額', '累計支出', '剩餘餘額']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 30
    
    # 範例資料（含公式）
    sample_data = [
        ['2026/01/15', '01-人事費', None, None, '補助款', 15000, None, None],
        ['2026/01/20', '02-業務費', None, None, '補助款', 25000, None, None],
        ['2026/02/10', '02-業務費', None, None, '自籌款', 18000, None, None],
        ['2026/02/28', '03-雜支', None, None, '補助款', 5000, None, None],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=6):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx == 4:  # 科目名稱 - VLOOKUP
                cell.value = f'=IFERROR(VLOOKUP(C{row_idx},計畫資料!$D$6:$E$20,1,FALSE)&" - "&VLOOKUP(C{row_idx},計畫資料!$D$6:$E$20,2,FALSE),"")'
            elif col_idx == 5:  # 核定金額 - VLOOKUP
                cell.value = f'=IFERROR(SUMIF(計畫資料!$D$6:$D$20,C{row_idx},計畫資料!$E$6:$E$20),"")'
                cell.number_format = '#,##0'
            elif col_idx == 8:  # 累計支出 - SUMIF
                cell.value = f'=SUMIF($C$6:$C$100,C{row_idx},$G$6:$G$100)'
                cell.number_format = '#,##0'
            elif col_idx == 9:  # 剩餘餘額 - 計算
                cell.value = f'=IFERROR(E{row_idx}-H{row_idx},"")'
                cell.number_format = '#,##0'
            else:
                cell.value = value
                if col_idx == 7:  # 本月報支金額
                    cell.number_format = '#,##0'
            
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN if col_idx in [2, 3, 6] else (RIGHT_ALIGN if col_idx in [5, 7, 8, 9] else LEFT_ALIGN)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL
    
    # 設定下拉選單 - 對應科目
    dv_budget = DataValidation(
        type="list",
        formula1='"01-人事費,02-業務費,03-雜支"',
        allow_blank=True
    )
    dv_budget.error = '請從下拉選單選擇有效的預算科目'
    dv_budget.errorTitle = '輸入錯誤'
    ws.add_data_validation(dv_budget)
    dv_budget.add('C6:C100')
    
    # 設定下拉選單 - 經費來源
    dv_source = DataValidation(
        type="list",
        formula1='"補助款,自籌款"',
        allow_blank=True
    )
    dv_source.error = '請選擇補助款或自籌款'
    dv_source.errorTitle = '輸入錯誤'
    ws.add_data_validation(dv_source)
    dv_source.add('F6:F100')
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # 公式說明區
    ws.merge_cells('B12:I12')
    ws['B12'] = '【VLOOKUP 公式說明】'
    ws['B12'].font = Font(name='Microsoft JhengHei', size=11, bold=True, color=THEME['accent'])
    
    formulas_info = [
        ('核定金額公式：', '=SUMIF(計畫資料!$D$6:$D$20,C6,計畫資料!$E$6:$E$20)'),
        ('累計支出公式：', '=SUMIF($C$6:$C$100,C6,$G$6:$G$100)'),
        ('剩餘餘額公式：', '=E6-H6 (核定金額 - 累計支出)'),
    ]
    
    for idx, (label, formula) in enumerate(formulas_info, start=13):
        ws.cell(row=idx, column=2, value=label).font = Font(name='Microsoft JhengHei', size=9, bold=True)
        ws.cell(row=idx, column=3, value=formula).font = Font(name='Consolas', size=9, color='7F8C8D')
    
    # 重要提醒
    ws.merge_cells('B17:I17')
    ws['B17'] = '【重要提醒】'
    ws['B17'].font = Font(name='Microsoft JhengHei', size=11, bold=True, color=THEME['danger'])
    
    ws.merge_cells('B18:I18')
    ws['B18'] = '• 「經費來源」欄位區分補助款與自籌款，系統可據此自動產出「經費撥付申請表」'
    ws['B18'].font = Font(name='Microsoft JhengHei', size=9, color=THEME['text'])
    
    ws.merge_cells('B19:I19')
    ws['B19'] = '• 人事費支出請注意文化部獎補助資訊網規範的上限（通常為核定金額的 30%）'
    ws['B19'].font = Font(name='Microsoft JhengHei', size=9, color=THEME['text'])


def main():
    """主程式"""
    wb = create_workbook()
    output_path = '/home/ubuntu/Mag/excel_template/原村管考系統_Excel匯入架構範本.xlsx'
    wb.save(output_path)
    print(f'Excel 範本已建立：{output_path}')
    return output_path


if __name__ == '__main__':
    main()
